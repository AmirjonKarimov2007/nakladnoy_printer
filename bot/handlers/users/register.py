from aiogram.types import *
from data.config import *
from loader import *
from states.admin_state import SuperAdminState
from filters.admins import IsAdmin

from openpyxl import load_workbook
import warnings

def check_user_id(user_id: int, file_path="informations/users.xlsx") -> bool:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active  # birinchi list
    
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if row[0] == user_id:
            return True
    return False




@dp.message_handler(commands='register',state='*')
async def register_user(message: Message):
    await message.answer(text=f"Iltimos {message.from_user.first_name},Smartup tomindan berilgan sotuvchilik idyingizni kiriting.")
    await SuperAdminState.SUPER_ADMIN_REGISTER_STATE.set()  
from aiogram.dispatcher import FSMContext


@dp.message_handler(content_types=ContentType.TEXT,state=SuperAdminState.SUPER_ADMIN_REGISTER_STATE)
async def user_regisration(message: Message,state:FSMContext):
    xabar = await message.answer('ID tekshiriliyapdi kuting...')
    saler_id = message.text
    if saler_id.isdigit():
        check= check_user_id(user_id=int(saler_id))
        if check:
            await db.update_saler_id(user_id=message.from_user.id,saler_id=int(saler_id))
            await xabar.delete()
            await message.answer("✅Siz muvaffaqiyatli ro'yxatdan o'tdingiz.")
            await state.finish()
        else:
           await message.answer("❌Siz notugri ID yubording. boshqattan urinib ko'ring. /register")
           await state.finish()
    else:
        await message.answer("ID faqat raqamlardan iborat bo'lishi kerak.")
