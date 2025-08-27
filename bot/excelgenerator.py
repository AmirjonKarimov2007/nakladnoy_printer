import win32com.client
import os
import asyncio
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image  # Yangi import: rasm qo'shish uchun
import os
import win32com.client
from functions import *
import pythoncom
import psutil
import os
import asyncio
import psutil
import pythoncom
import win32com.client
from data.config import USD
from loader import db
import win32print

def set_default_printer(printer_name="Canon MF3010"):
    try:
        win32print.SetDefaultPrinter(printer_name)
        print(f"Default printer set to: {printer_name}")
    except Exception as e:
        print(f"Error: {e}")
def kill_task():
    task_name = 'EXCEL.EXE'
    for proc in psutil.process_iter(['pid', 'name']):
        if task_name.lower() in proc.info['name'].lower(): 
            try:
                proc.kill()  
            except psutil.NoSuchProcess:
                pass
            except psutil.AccessDenied:
                pass
def print_excel_file_sync(file_path):
    try:
        set_default_printer()
        pythoncom.CoInitialize()
        abs_path = os.path.abspath(file_path)  
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False 
        wb = excel.Workbooks.Open(abs_path)  
        ws = wb.ActiveSheet

        # Chop etish sozlamalari
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False 
        ws.PageSetup.PaperSize = 9  # A4 qog'oz o'lchami
        ws.PageSetup.Orientation = 1  # Portret rejimi

        # Chop etish (faqat qora-oq rejimda)
        # ws.PageSetup.PrintInBlackAndWhite = True  # <-- Qora-oq rejim
        wb.PrintOut()
        
        # Faylni yopish va Excel ilovasini to'xtatish
        wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()
        return True
    except Exception as e:
        print(f"Xatolik yuz berdi: {e}")
        return False 
    finally:
        kill_task()

async def print_excel_file(file_path):
    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, print_excel_file_sync, file_path)
        if result:
            return True
        else:
            return False
    except Exception as e:
        print(f"Asinxron xatolik yuz berdi: {e}")
        return False
    



import qrcode
import os

def qrcode_generator(deal_id):
    url = f"https://hackingbot.alwaysdata.net/?order_id={deal_id}"
    
    qr = qrcode.make(url)
    
    filename = f"qrcodes/{deal_id}.png"
    
    save_path = os.path.join(os.getcwd(), filename)
    qr.save(save_path)
    
    return save_path


def get_box_details(quantity, box_quant):
    if not isinstance(box_quant, int):
        return f"{quantity} шт."
    quantity = int(quantity)
    box_quant = int(box_quant)
    if box_quant == 0:
        return "0"
    full_boxes = quantity // box_quant
    remaining_items = quantity % box_quant
    if remaining_items == 0:
        return f"{full_boxes} коробка"
    elif full_boxes == 0:
        return f"{remaining_items} шт."
    else:
        return f"{full_boxes} коробка, {remaining_items} шт."
def format_barcode(barcode):
    barcode = str(barcode)
    if '|' in barcode:
        return barcode.replace('|', '\n')
    return barcode
def format_product_name(name, max_length=45):
    formatted_name = ' '.join(name.split())
    if len(formatted_name) > max_length:
        formatted_name = '\n'.join([formatted_name[i:i+max_length] for i in range(0, len(formatted_name), max_length)])
    return formatted_name
def set_column_width(ws, col_idx, start_row, end_row):
    max_length = 0
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width



from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import XDRPositiveSize2D

def add_qrcode_centered(ws, qrcode_path, cell_address='C2'):
    """
    QR-kodni berilgan katak (masalan, 'C3') gorizontal markaziga joylashtiradi
    """
    img = Image(qrcode_path)
    img.width = 130  # px
    img.height = 130

    # Ustun harfi va qator raqamini ajratib olamiz
    col_letter = ''.join(filter(str.isalpha, cell_address))
    row_number = int(''.join(filter(str.isdigit, cell_address)))
    col_index = column_index_from_string(col_letter) - 1  # 0-based

    # Ustun kengligi (agar yo‘q bo‘lsa, default 8.43)
    col_width = ws.column_dimensions[col_letter].width or 8.43
    col_width_px = col_width * 7.5  # taxminan 1 unit ≈ 7.5 px

    # Gorizontal offset: katak markaziga joylash
    offset_x_px = max((col_width_px - img.width) / 2, 0)
    offset_x_emu = pixels_to_EMU(offset_x_px)

    # Vertikal offset (katak ichida yuqoridan boshlaydi — 0 bo‘lishi mumkin)
    offset_y_emu = pixels_to_EMU(0)

    # Anchor sozlash (qayerga joylash)
    marker = AnchorMarker(col=col_index,
                          colOff=offset_x_emu,
                          row=row_number - 1,
                          rowOff=offset_y_emu)

    img.anchor = OneCellAnchor(_from=marker,
                               ext=XDRPositiveSize2D(pixels_to_EMU(img.width),
                                                  pixels_to_EMU(img.height)))

    ws.add_image(img)
    return cell_address


async def process_order(deal_id, output_path, moment,moneytype = "usd"):
    try:
        if moment == 'today':
            order_info = await get_today_order_info_by_deal_id(deal_id)
        elif moment == 'yesterday':
            order_info = await get_lastday_order_info_by_deal_id(deal_id)
        else:
            order_info = await get_order_info_by_deal_id(deal_id)

        products = order_info['order_products']
        currency_code = order_info['currency_code']
        all_price = order_info['total_amount']
        date = order_info['booked_date']
        saler_id = order_info['sales_manager_id']
        time = order_info['deal_time']
        name = order_info['sales_manager_name']
        room_name = order_info['room_name']
        
        client_name = order_info['person_name']

        wb = load_workbook("shablon.xlsx")
        ws = wb.active
        user = await db.select_user(saler_id=int(saler_id))
        if user:
            phone_number = user[0]['phone_number']
        else:
            phone_number = " "
        qrcode_path = qrcode_generator(deal_id)  # O'zgartirish: "path" o'rniga "qrcode_path" deb o'zgartirdim, chunki keyinroq ishlatiladi
        ws['A4'] = f"Дата отгрузки: {date}"  
        ws['A8'] = f"Время заказа: {time}" 
        ws['D2'] = f"Торговый представитель: {name}" 
        
        ws['D4'] = f"Контрагент: {client_name}" 
        ws['D3'] = f"Телефон торгового представителя: +{phone_number}" 

        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        total_count = 0
        row = 11
        E_qator = 0
        warehouse = 0
        for i, product in enumerate(products, start=1):
            ws[f'A{row}'] = i
            barcode = format_barcode(product['product_barcode'])
            ws[f'B{row}'] = barcode
            product_name = format_product_name(product['product_name'])
            ws[f'C{row}'] = product_name
            ws[f'D{row}'] = f"{product['order_quant']}"
            box_quant = get_box_details(quantity=product['order_quant'], box_quant=product['box_quant'])
            warehouse = product['warehouse_code']
            ws[f'E{row}'] = box_quant
            product_price = float(product['product_price'])
          
            if product_price.is_integer():
                ws[f'F{row}'] = f"{int(product_price):,}"
            else:
                ws[f'F{row}'] = f"{product_price:,.2f}"
            
            sold_amount = float(product['sold_amount'])
            if sold_amount.is_integer():
                ws[f'G{row}'] = f"{int(sold_amount):,}"
            else:
                ws[f'G{row}'] = f"{sold_amount:,.2f}"

            ws.row_dimensions[row].auto_size = True 
            total_count += int(product['order_quant'])
            box_quant_length = len(str(box_quant))
            if box_quant_length > E_qator:
                E_qator = box_quant_length

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = border_style
                cell.font = Font(size=12) 
                if col in [4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            row += 1

        ws.row_dimensions[row].auto_size = True 

        for col in range(1, 8):
            if col == 3:
                ws.column_dimensions[get_column_letter(col)].width = 50
            else:
                set_column_width(ws, col, start_row=12, end_row=row-1)

        ws.column_dimensions['B'].width = 16
        black_row = row
        black_fill = PatternFill(start_color="D5D5D5", end_color="D5D5D5", fill_type="solid")

        for col in range(1, 8):
            ws.cell(row=black_row, column=col).fill = black_fill
            cell = ws.cell(row=black_row, column=col)
            cell.border = border_style

        ws.merge_cells(f'A{black_row}:G{black_row}')

        if currency_code == "840":
            uzs = f"UZS: {float(all_price) * USD:,.2f}"
            ws[f'A{black_row}'] = f"Итого: {all_price}$|->{uzs}"
        else:
            ws[f'A{black_row}'] = f"Итого: {all_price}"
        if warehouse=="111":
            ws['C10'] = ""  # Falcon nomi o'chirildi, chunki QR o'rniga joylashadi
        elif warehouse=="11":
            ws['C10'] = ""  # Xuddi shunday

        ws[f'A{black_row}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        ws[f'A{black_row}'].font = Font(bold=True)
        ws[f'A{black_row}'].border = border_style

        ws['D10'] = total_count
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].font = Font(bold=True)

        ws['G10'] = all_price
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G10'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = E_qator + 3
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 12

        position = add_qrcode_centered(ws, qrcode_path)


        # Faylni saqlash
        if not os.path.exists('orders'):
            os.makedirs('orders')
        wb.save(f"orders/{output_path}.xlsx")
        return True
    except Exception as e:
        if "Permission denied" in str(e):
            kill_task()
            return True
        else:
            print(f"Xatolik yuz beribdi: {e}")
            return False
async def main():
    await process_order(deal_id="120462256",output_path="1",moment='today')
def update_excel_with_products(file_path: str, products_data: list, output_path: str = "tahrirlangan_rasxod.xlsx"):
    wb = load_workbook(file_path)
    ws = wb.active
    start_row = 2
    for index, product in enumerate(products_data, start=start_row):
        ws.cell(row=index, column=1, value=product.get("product_code", ""))
        ws.cell(row=index, column=2, value=product.get("order_quant", 0))
        ws.cell(row=index, column=3, value="Y")
    wb.save(output_path)
    return output_path